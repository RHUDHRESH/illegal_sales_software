"""
Lead Export Functionality - CSV, Excel, JSON, PDF
"""
import logging
import csv
import json
from io import BytesIO, StringIO
from typing import List, Dict, Any, Optional
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from sqlalchemy.orm import Session
from ..database import Lead, Company, Contact

logger = logging.getLogger(__name__)


class LeadExporter:
    """
    Export leads in various formats
    """

    def __init__(self):
        self.export_history = []

    def export_to_csv(
        self,
        leads: List[Lead],
        include_dossier: bool = False
    ) -> str:
        """
        Export leads to CSV format

        Args:
            leads: List of Lead objects
            include_dossier: Whether to include full dossier (makes CSV large)

        Returns:
            CSV string
        """
        output = StringIO()
        writer = csv.writer(output)

        # Headers
        headers = [
            "ID",
            "Company Name",
            "Website",
            "Total Score",
            "Score Bucket",
            "ICP Fit Score",
            "Marketing Pain Score",
            "Data Quality Score",
            "Status",
            "Role Type",
            "Pain Tags",
            "Situation",
            "Problem",
            "Key Pain",
            "Economic Buyer",
            "Chaos Flags",
            "Contact Emails",
            "Contact Phones",
            "Created At"
        ]

        if include_dossier:
            headers.extend([
                "Snapshot",
                "Uncomfortable Truth",
                "Reframe Suggestion",
                "Challenger Insight"
            ])

        writer.writerow(headers)

        # Data rows
        for lead in leads:
            row = [
                lead.id,
                lead.company.name if lead.company else "",
                lead.company.website if lead.company else "",
                lead.total_score,
                lead.score_bucket,
                lead.score_icp_fit,
                lead.score_marketing_pain,
                lead.score_data_quality,
                lead.status,
                lead.role_type,
                ", ".join(lead.pain_tags) if lead.pain_tags else "",
                lead.situation or "",
                lead.problem or "",
                lead.key_pain or "",
                lead.economic_buyer_guess or "",
                ", ".join(lead.chaos_flags) if lead.chaos_flags else "",
                self._get_contact_emails(lead),
                self._get_contact_phones(lead),
                lead.created_at.isoformat() if lead.created_at else ""
            ]

            if include_dossier:
                row.extend([
                    self._extract_dossier_field(lead.context_dossier, "snapshot"),
                    self._extract_dossier_field(lead.context_dossier, "uncomfortable_truth"),
                    lead.reframe_suggestion or "",
                    lead.challenger_insight or ""
                ])

            writer.writerow(row)

        self._log_export("CSV", len(leads))
        return output.getvalue()

    def export_to_excel(
        self,
        leads: List[Lead],
        include_dossier: bool = False,
        include_summary: bool = True
    ) -> BytesIO:
        """
        Export leads to Excel format with formatting

        Args:
            leads: List of Lead objects
            include_dossier: Whether to include full dossier
            include_summary: Whether to include summary sheet

        Returns:
            BytesIO object containing Excel file
        """
        output = BytesIO()
        workbook = Workbook()

        # Remove default sheet
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        # Summary sheet
        if include_summary:
            summary_sheet = workbook.create_sheet("Summary")
            self._create_summary_sheet(summary_sheet, leads)

        # Leads sheet
        leads_sheet = workbook.create_sheet("Leads")
        self._create_leads_sheet(leads_sheet, leads, include_dossier)

        # Hot Leads sheet (score >= 80)
        hot_leads = [l for l in leads if l.total_score and l.total_score >= 80]
        if hot_leads:
            hot_sheet = workbook.create_sheet("Hot Leads")
            self._create_leads_sheet(hot_sheet, hot_leads, include_dossier=True)

        workbook.save(output)
        output.seek(0)

        self._log_export("Excel", len(leads))
        return output

    def export_to_json(
        self,
        leads: List[Lead],
        include_dossier: bool = True
    ) -> str:
        """
        Export leads to JSON format

        Args:
            leads: List of Lead objects
            include_dossier: Whether to include full dossier

        Returns:
            JSON string
        """
        leads_data = []

        for lead in leads:
            lead_dict = {
                "id": lead.id,
                "company": {
                    "id": lead.company.id if lead.company else None,
                    "name": lead.company.name if lead.company else None,
                    "website": lead.company.website if lead.company else None,
                },
                "scores": {
                    "total": lead.total_score,
                    "bucket": lead.score_bucket,
                    "icp_fit": lead.score_icp_fit,
                    "marketing_pain": lead.score_marketing_pain,
                    "data_quality": lead.score_data_quality
                },
                "classification": {
                    "role_type": lead.role_type,
                    "pain_tags": lead.pain_tags,
                    "chaos_flags": lead.chaos_flags,
                    "economic_buyer": lead.economic_buyer_guess
                },
                "spin": {
                    "situation": lead.situation,
                    "problem": lead.problem,
                    "implication": lead.implication,
                    "need_payoff": lead.need_payoff
                },
                "key_insights": {
                    "key_pain": lead.key_pain,
                    "silver_bullet_phrases": lead.silver_bullet_phrases
                },
                "contacts": {
                    "emails": self._get_contact_emails_list(lead),
                    "phones": self._get_contact_phones_list(lead)
                },
                "status": lead.status,
                "owner": lead.owner,
                "created_at": lead.created_at.isoformat() if lead.created_at else None,
                "updated_at": lead.updated_at.isoformat() if lead.updated_at else None
            }

            if include_dossier and lead.context_dossier:
                lead_dict["dossier"] = {
                    "snapshot": self._extract_dossier_field(lead.context_dossier, "snapshot"),
                    "uncomfortable_truth": self._extract_dossier_field(lead.context_dossier, "uncomfortable_truth"),
                    "reframe_suggestion": lead.reframe_suggestion,
                    "challenger_insight": lead.challenger_insight
                }

            leads_data.append(lead_dict)

        self._log_export("JSON", len(leads))
        return json.dumps(leads_data, indent=2)

    def export_to_pdf(
        self,
        leads: List[Lead],
        include_dossier: bool = True,
        title: str = "Lead Report"
    ) -> BytesIO:
        """
        Export leads to PDF format

        Args:
            leads: List of Lead objects
            include_dossier: Whether to include full dossier
            title: Report title

        Returns:
            BytesIO object containing PDF file
        """
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1  # Center
        )
        elements.append(Paragraph(title, title_style))

        # Summary stats
        total_leads = len(leads)
        hot_leads = len([l for l in leads if l.total_score and l.total_score >= 80])
        warm_leads = len([l for l in leads if l.total_score and 60 <= l.total_score < 80])

        summary_text = f"""
        <b>Total Leads:</b> {total_leads}<br/>
        <b>Hot Leads (80+):</b> {hot_leads}<br/>
        <b>Warm Leads (60-79):</b> {warm_leads}<br/>
        <b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}<br/>
        """
        elements.append(Paragraph(summary_text, styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))

        # Leads table
        if include_dossier:
            # Detailed view with dossier
            for i, lead in enumerate(leads[:50]):  # Limit to 50 for PDF size
                elements.append(self._create_lead_pdf_section(lead, styles))
                if i < len(leads) - 1:
                    elements.append(Spacer(1, 0.2*inch))
                if (i + 1) % 3 == 0:  # Page break every 3 leads
                    elements.append(PageBreak())
        else:
            # Summary table view
            table_data = [
                ["Company", "Score", "Bucket", "Pain", "Status"]
            ]

            for lead in leads:
                table_data.append([
                    lead.company.name if lead.company else "",
                    str(lead.total_score or 0),
                    lead.score_bucket or "",
                    (lead.key_pain or "")[:30] + "..." if lead.key_pain and len(lead.key_pain) > 30 else lead.key_pain or "",
                    lead.status or ""
                ])

            table = Table(table_data, colWidths=[2*inch, 0.7*inch, 0.9*inch, 2.5*inch, 0.9*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            elements.append(table)

        doc.build(elements)
        output.seek(0)

        self._log_export("PDF", len(leads))
        return output

    def _create_summary_sheet(self, sheet, leads: List[Lead]):
        """Create Excel summary sheet"""
        # Title
        sheet["A1"] = "Lead Report Summary"
        sheet["A1"].font = Font(size=16, bold=True)

        # Stats
        sheet["A3"] = "Total Leads:"
        sheet["B3"] = len(leads)

        sheet["A4"] = "Hot Leads (80+):"
        sheet["B4"] = len([l for l in leads if l.total_score and l.total_score >= 80])

        sheet["A5"] = "Warm Leads (60-79):"
        sheet["B5"] = len([l for l in leads if l.total_score and 60 <= l.total_score < 80])

        sheet["A6"] = "Nurture Leads (40-59):"
        sheet["B6"] = len([l for l in leads if l.total_score and 40 <= l.total_score < 60])

        sheet["A7"] = "Generated:"
        sheet["B7"] = datetime.now().strftime('%Y-%m-%d %H:%M')

        # Average scores
        sheet["A9"] = "Average Total Score:"
        avg_score = sum(l.total_score or 0 for l in leads) / len(leads) if leads else 0
        sheet["B9"] = round(avg_score, 2)

        # Format
        for row in range(3, 10):
            sheet[f"A{row}"].font = Font(bold=True)

    def _create_leads_sheet(self, sheet, leads: List[Lead], include_dossier: bool):
        """Create Excel leads sheet with data"""
        # Headers
        headers = [
            "ID", "Company", "Website", "Total Score", "Bucket",
            "ICP Fit", "Pain Score", "Status", "Role Type",
            "Key Pain", "Economic Buyer", "Emails", "Phones"
        ]

        if include_dossier:
            headers.extend(["Snapshot", "Uncomfortable Truth", "Challenger Insight"])

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data
        for row_idx, lead in enumerate(leads, 2):
            data = [
                lead.id,
                lead.company.name if lead.company else "",
                lead.company.website if lead.company else "",
                lead.total_score or 0,
                lead.score_bucket or "",
                lead.score_icp_fit or 0,
                lead.score_marketing_pain or 0,
                lead.status or "",
                lead.role_type or "",
                lead.key_pain or "",
                lead.economic_buyer_guess or "",
                self._get_contact_emails(lead),
                self._get_contact_phones(lead)
            ]

            if include_dossier:
                data.extend([
                    self._extract_dossier_field(lead.context_dossier, "snapshot"),
                    self._extract_dossier_field(lead.context_dossier, "uncomfortable_truth"),
                    lead.challenger_insight or ""
                ])

            for col, value in enumerate(data, 1):
                cell = sheet.cell(row=row_idx, column=col, value=value)

                # Color code by score
                if col == 4 and lead.total_score:  # Total Score column
                    if lead.total_score >= 80:
                        cell.fill = PatternFill(start_color="FFE7E7", end_color="FFE7E7", fill_type="solid")
                    elif lead.total_score >= 60:
                        cell.fill = PatternFill(start_color="FFFEE7", end_color="FFFEE7", fill_type="solid")

        # Auto-size columns
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    def _create_lead_pdf_section(self, lead: Lead, styles) -> List:
        """Create PDF section for a single lead"""
        elements = []

        # Lead header
        company_name = lead.company.name if lead.company else "Unknown Company"
        score_text = f" (Score: {lead.total_score or 0})"

        header = Paragraph(
            f"<b>{company_name}</b>{score_text}",
            styles['Heading2']
        )
        elements.append(header)

        # Key info
        info_text = f"""
        <b>Score Bucket:</b> {lead.score_bucket or 'N/A'}<br/>
        <b>Status:</b> {lead.status or 'new'}<br/>
        <b>Key Pain:</b> {lead.key_pain or 'N/A'}<br/>
        <b>Economic Buyer:</b> {lead.economic_buyer_guess or 'Unknown'}<br/>
        """

        if lead.company and lead.company.website:
            info_text += f"<b>Website:</b> {lead.company.website}<br/>"

        elements.append(Paragraph(info_text, styles['Normal']))

        # Dossier
        if lead.context_dossier:
            snapshot = self._extract_dossier_field(lead.context_dossier, "snapshot")
            if snapshot:
                elements.append(Paragraph(f"<b>Snapshot:</b> {snapshot}", styles['Normal']))

            insight = lead.challenger_insight
            if insight:
                elements.append(Paragraph(f"<b>Lead With:</b> {insight}", styles['Normal']))

        return elements

    def _get_contact_emails(self, lead: Lead) -> str:
        """Get comma-separated list of contact emails"""
        emails = self._get_contact_emails_list(lead)
        return ", ".join(emails)

    def _get_contact_emails_list(self, lead: Lead) -> List[str]:
        """Get list of contact emails"""
        if not lead.company or not lead.company.contacts:
            return []
        return [c.email for c in lead.company.contacts if c.email]

    def _get_contact_phones(self, lead: Lead) -> str:
        """Get comma-separated list of contact phones"""
        phones = self._get_contact_phones_list(lead)
        return ", ".join(phones)

    def _get_contact_phones_list(self, lead: Lead) -> List[str]:
        """Get list of contact phones"""
        if not lead.company or not lead.company.contacts:
            return []
        return [c.phone for c in lead.company.contacts if c.phone]

    def _extract_dossier_field(self, dossier: Optional[str], field: str) -> str:
        """Extract field from dossier JSON"""
        if not dossier:
            return ""
        try:
            dossier_dict = json.loads(dossier) if isinstance(dossier, str) else dossier
            return dossier_dict.get(field, "")
        except:
            return ""

    def _log_export(self, format_type: str, count: int):
        """Log export operation"""
        self.export_history.append({
            "format": format_type,
            "count": count,
            "timestamp": datetime.now().isoformat()
        })

        # Keep only last 100 exports
        if len(self.export_history) > 100:
            self.export_history = self.export_history[-100:]


# Global instance
lead_exporter = LeadExporter()
